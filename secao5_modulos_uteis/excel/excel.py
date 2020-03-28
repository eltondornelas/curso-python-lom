"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from random import uniform

"""

# utilizando a planilha existente
pedidos = openpyxl.load_workbook('pedidos.xlsx')  # joga a planilha para a variável
nome_planilhas = pedidos.sheetnames  # quais as planilhas (abas) que estão dentro do arquivo 
planilha1 = pedidos['Página1']
# print(planilha['b4'].value)

# for linha in planilha1['a1:c2']
for linha in range(5, 16):
    # print(linha.value)
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
"""

# criando uma planilha
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

# planilha1['B3'].value = 2200

for linha in range(1, 11):
    # print(len(linha))  # quantas colunas tem nessa linha, estranho que era pra ser 3 o retorno
    # print(linha[0].value)  # coluna A
    # if linha[0].value is not None:
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Elton {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Amanda {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Thor {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha.xlsx')

